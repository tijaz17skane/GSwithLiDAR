#!/usr/bin/env python3
"""
Build an Excel workbook of experimental results with a custom table layout.

- Scans a root directory for results_test.json and results_train.json
- Parses experiment names into (Initial PC Source, Method, Condition)
- Rounds metrics to 4 decimals
- Writes one sheet per Metric (PSNR, SSIM, LPIPS) and per result type (test, train)
  with the following column layout:
    [Initial PC Source] | [Base Method] | [DepthRegularization Method] | [separator] | [Base Method (With Scaffold)] | [Depth Regularization Method (With Scaffold)]
- The first row merges B1:F1 with the metric name as a header

Usage:
    python report_all_resultsFormatted.py --dir <root_dir> --out_xlsx <path_to_excel>
    Optional:
        --out_xlsx_tt <path>        Create a workbook with two sheets (Test, Train), stacking PSNR/SSIM/LPIPS tables
        --out_xlsx_metrics <path>   Create a workbook with one sheet per metric, each containing Train then Test sections
        --out_relevant <path>       Create filtered versions (relevant categories only) of all three outputs.
                                                                If <path> ends with .xlsx, it is used as the base and _TestTrain/_ByMetric are derived.
                                                                If <path> is a directory, files relevant.xlsx, relevant_TestTrain.xlsx, relevant_ByMetric.xlsx are created.
"""

from pathlib import Path
import os
import json
import argparse
import re
from typing import Dict, List, Any, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill
except Exception as e:
    openpyxl = None


def find_results_files(root_dir: str) -> Dict[str, List[Path]]:
    results = {"test": [], "train": []}
    root = Path(root_dir)
    for fp in root.rglob('results_test.json'):
        results['test'].append(fp)
    for fp in root.rglob('results_train.json'):
        results['train'].append(fp)
    return results


def extract_experiment_name(filepath: Path) -> str:
    return filepath.parent.name


def parse_experiment_name(experiment_name: str) -> Tuple[str, str, str]:
    """
    Parse experiment_name into (Initial PC Source, Method, Condition).
    - Method: 'base' -> 'Base'; 'depthReg'/'DepthRegularization' -> 'DepthRegularization'
    - Condition: contains 'wScaff' -> 'With Scaffold'; else 'N/A'
    - Initial PC Source: remove leading 'section...' and method/condition tokens
    """
    s = experiment_name or ''
    s = re.sub(r'^section[_\d]*', '', s, flags=re.IGNORECASE)

    condition = 'N/A'
    if 'wScaff' in s:
        condition = 'With Scaffold'
        s = s.replace('wScaff', '')

    method = 'Base'
    if re.search(r'depthreg|depthReg|DepthReg|DepthRegularization', experiment_name, flags=re.IGNORECASE):
        method = 'DepthRegularization'
        s = re.sub(r'(_?depthreg|_?depthReg|_?DepthReg|_?DepthRegularization)', '', s)
    elif re.search(r'\bbase\b|base$|Base$', experiment_name):
        method = 'Base'
        s = re.sub(r'(_?base|Base)$', '', s)

    s = s.strip('_').replace('_', '')
    initial_pc = s
    return initial_pc, method, condition


def load_records_from_file(filepath: Path, result_type: str) -> List[Dict[str, Any]]:
    experiment_name = extract_experiment_name(filepath)
    out: List[Dict[str, Any]] = []
    with open(filepath, 'r') as f:
        data = json.load(f)
    for method_name, metrics in data.items():
        rec: Dict[str, Any] = {
            'experiment_name': experiment_name,
            'result_type': result_type,
            'method_raw': method_name,
            'file_path': str(filepath),
        }
        if isinstance(metrics, dict):
            for k, v in metrics.items():
                if isinstance(v, (int, float)):
                    rec[k] = round(float(v), 4)
                else:
                    rec[k] = v
        out.append(rec)
    return out


def collect_records(root_dir: str) -> List[Dict[str, Any]]:
    files = find_results_files(root_dir)
    records: List[Dict[str, Any]] = []
    # preserve discovery order
    for fp in files['test']:
        records.extend(load_records_from_file(fp, 'test'))
    for fp in files['train']:
        records.extend(load_records_from_file(fp, 'train'))
    return records


def _parse_rows(records: List[Dict[str, Any]], metrics: Tuple[str, ...]) -> List[Dict[str, Any]]:
    """Normalize records into row dicts for easier table construction."""
    parsed_rows: List[Dict[str, Any]] = []
    for r in records:
        initial, parsed_method, condition = parse_experiment_name(r.get('experiment_name', ''))
        for metric in metrics:
            if metric in r:
                val = r.get(metric)
                try:
                    val = round(float(val), 4)
                except Exception:
                    pass
                parsed_rows.append({
                    'result_type': r.get('result_type'),
                    'initial': initial,
                    'method': parsed_method,
                    'condition': condition,
                    'metric': metric,
                    'value': val,
                })
    return parsed_rows


def _coerce_number(v):
    """Return a float if v is numeric-or-convertible, otherwise None (so Excel cell is blank)."""
    if v is None:
        return None
    if v == '':
        return None
    try:
        return float(v)
    except Exception:
        return None


def build_excel(records: List[Dict[str, Any]], out_xlsx: str):
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to output Excel. Please install it: pip install openpyxl")

    metrics: Tuple[str, ...] = ('PSNR', 'SSIM', 'LPIPS')  # user asked separate tables for PSNR, SSIM, LPIPS

    # Parse records into normalized rows with parsed fields
    parsed_rows = _parse_rows(records, metrics)

    wb = openpyxl.Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # Build one sheet per metric and result_type
    for metric in metrics:
        for result_type in ('test', 'train'):
            sheet_name = f"{metric}_{result_type}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Collect initials in discovery order
            initials: List[str] = []
            for pr in parsed_rows:
                if pr['result_type'] != result_type or pr['metric'] != metric:
                    continue
                if pr['initial'] not in initials:
                    initials.append(pr['initial'])

            # Determine presence of any 'With Scaffold' entries for this sheet
            has_with_scaffold = any(
                pr['result_type'] == result_type and pr['metric'] == metric and pr['condition'] == 'With Scaffold'
                for pr in parsed_rows
            )

            # Header row: merged metric name across B1:Fn depending on layout
            if initials:
                end_col = 6 if has_with_scaffold else 3
                ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=end_col)
                ws.cell(row=1, column=2, value=metric)
                ws.cell(row=1, column=2).alignment = Alignment(horizontal='center')

            # Column headers row
            if has_with_scaffold:
                headers = [
                    'Initial PC Source',
                    'Base Method',
                    'DepthRegularization Method',
                    'With Scaffold',  # vertically merged across data rows
                    'Base Method',
                    'Depth Regularization Method'
                ]
            else:
                headers = [
                    'Initial PC Source',
                    'Base Method',
                    'DepthRegularization Method'
                ]
            for c, h in enumerate(headers, start=1):
                ws.cell(row=2, column=c, value=h)
                ws.cell(row=2, column=c).alignment = Alignment(horizontal='center')

            # Build lookup for quick value access
            lookup: Dict[Tuple[str, str, str], Any] = {}
            for pr in parsed_rows:
                if pr['result_type'] != result_type or pr['metric'] != metric:
                    continue
                lookup[(pr['initial'], pr['method'], pr['condition'])] = pr['value']

            # Fill table rows
            r_idx = 3
            for initial in initials:
                ws.cell(row=r_idx, column=1, value=initial)
                # Without Scaffold
                v2 = _coerce_number(lookup.get((initial, 'Base', 'N/A'), None))
                v3 = _coerce_number(lookup.get((initial, 'DepthRegularization', 'N/A'), None))
                ws.cell(row=r_idx, column=2, value=v2)
                ws.cell(row=r_idx, column=3, value=v3)
                if has_with_scaffold:
                    # With Scaffold columns (coerce to numbers)
                    v5 = _coerce_number(lookup.get((initial, 'Base', 'With Scaffold'), None))
                    v6 = _coerce_number(lookup.get((initial, 'DepthRegularization', 'With Scaffold'), None))
                    ws.cell(row=r_idx, column=5, value=v5)
                    ws.cell(row=r_idx, column=6, value=v6)
                r_idx += 1

            # If layout has With Scaffold, merge Column 4 from header row 2 through last data row
            if has_with_scaffold and r_idx > 3:
                ws.merge_cells(start_row=2, start_column=4, end_row=r_idx-1, end_column=4)
                ws.cell(row=2, column=4, value='With Scaffold')
                ws.cell(row=2, column=4).alignment = Alignment(horizontal='center', vertical='center')

            # Column widths
            max_col = 6 if has_with_scaffold else 3
            for i in range(1, max_col + 1):
                ws.column_dimensions[get_column_letter(i)].width = 26

            # Highlight best and second best values across metric columns
            value_cells = []
            # rows start at 3 to r_idx-1
            for rr in range(3, r_idx):
                cols = (2, 3, 5, 6) if has_with_scaffold else (2, 3)
                for cc in cols:
                    cell = ws.cell(row=rr, column=cc)
                    try:
                        val = float(cell.value)
                        value_cells.append((val, cell))
                    except (TypeError, ValueError):
                        pass

            if value_cells:
                # Sort depending on metric: PSNR/SSIM higher is better, LPIPS lower is better
                reverse_order = metric in ('PSNR', 'SSIM')
                sorted_vals = sorted(value_cells, key=lambda x: x[0], reverse=reverse_order)
                # For LPIPS smallest first; for PSNR/SSIM largest first (due to reverse flag)
                best_val = sorted_vals[0][0]
                # Find second distinct value
                second_val = None
                for v, c in sorted_vals[1:]:
                    if v != best_val:
                        second_val = v
                        break
                # Define fills
                best_fill = PatternFill(start_color='32B61B', end_color='32B61B', fill_type='solid')
                second_fill = PatternFill(start_color='EDB64E', end_color='EDB64E', fill_type='solid')
                for v, c in value_cells:
                    if v == best_val:
                        c.fill = best_fill
                    elif second_val is not None and v == second_val:
                        c.fill = second_fill

    # Ensure output directory exists
    out_dir = os.path.dirname(out_xlsx)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_xlsx)
    print(f"Saved Excel workbook to: {out_xlsx}")


def _write_metric_section(ws, parsed_rows: List[Dict[str, Any]], metric: str, result_type: str, start_row: int) -> int:
    """Write one metric table (same layout as individual sheets) at start_row; return next row index."""
    # Collect initials in discovery order
    initials: List[str] = []
    for pr in parsed_rows:
        if pr['result_type'] != result_type or pr['metric'] != metric:
            continue
        if pr['initial'] not in initials:
            initials.append(pr['initial'])

    # Determine presence of any 'With Scaffold' entries for this block
    has_with_scaffold = any(
        pr['result_type'] == result_type and pr['metric'] == metric and pr['condition'] == 'With Scaffold'
        for pr in parsed_rows
    )

    row = start_row
    if initials:
        end_col = 6 if has_with_scaffold else 3
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end_col)
        ws.cell(row=row, column=2, value=f"{metric} ({result_type})")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    row += 1

    # Headers
    if has_with_scaffold:
        headers = [
            'Initial PC Source',
            'Base Method',
            'DepthRegularization Method',
            'With Scaffold',
            'Base Method',
            'Depth Regularization Method'
        ]
    else:
        headers = [
            'Initial PC Source',
            'Base Method',
            'DepthRegularization Method'
        ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
        ws.cell(row=row, column=c).alignment = Alignment(horizontal='center')

    # Build lookup and fill
    lookup: Dict[Tuple[str, str, str], Any] = {}
    for pr in parsed_rows:
        if pr['result_type'] != result_type or pr['metric'] != metric:
            continue
        lookup[(pr['initial'], pr['method'], pr['condition'])] = pr['value']

    r_idx = row + 1
    for initial in initials:
        ws.cell(row=r_idx, column=1, value=initial)
        ws.cell(row=r_idx, column=2, value=_coerce_number(lookup.get((initial, 'Base', 'N/A'), None)))
        ws.cell(row=r_idx, column=3, value=_coerce_number(lookup.get((initial, 'DepthRegularization', 'N/A'), None)))
        if has_with_scaffold:
            ws.cell(row=r_idx, column=5, value=_coerce_number(lookup.get((initial, 'Base', 'With Scaffold'), None)))
            ws.cell(row=r_idx, column=6, value=_coerce_number(lookup.get((initial, 'DepthRegularization', 'With Scaffold'), None)))
        r_idx += 1

    # Merge 'With Scaffold' col
    if has_with_scaffold and r_idx > row + 1:
        ws.merge_cells(start_row=row, start_column=4, end_row=r_idx-1, end_column=4)
        ws.cell(row=row, column=4, value='With Scaffold')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')

    # Highlight best/second-best
    value_cells = []
    for rr in range(row+1, r_idx):
        cols = (2, 3, 5, 6) if has_with_scaffold else (2, 3)
        for cc in cols:
            cell = ws.cell(row=rr, column=cc)
            try:
                val = float(cell.value)
                value_cells.append((val, cell))
            except (TypeError, ValueError):
                pass
    if value_cells:
        reverse_order = metric in ('PSNR', 'SSIM')
        sorted_vals = sorted(value_cells, key=lambda x: x[0], reverse=reverse_order)
        best_val = sorted_vals[0][0]
        second_val = None
        for v, c in sorted_vals[1:]:
            if v != best_val:
                second_val = v
                break
        best_fill = PatternFill(start_color='32B61B', end_color='32B61B', fill_type='solid')
        second_fill = PatternFill(start_color='EDB64E', end_color='EDB64E', fill_type='solid')
        for v, c in value_cells:
            if v == best_val:
                c.fill = best_fill
            elif second_val is not None and v == second_val:
                c.fill = second_fill

    # Return next empty row + one spacer
    return r_idx + 1


def build_excel_test_train(records: List[Dict[str, Any]], out_xlsx_tt: str):
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to output Excel. Please install it: pip install openpyxl")

    metrics: Tuple[str, ...] = ('PSNR', 'SSIM', 'LPIPS')
    parsed_rows = _parse_rows(records, metrics)

    wb = openpyxl.Workbook()
    ws_test = wb.active
    ws_test.title = 'Test'
    ws_train = wb.create_sheet(title='Train')

    # Set column widths for both sheets
    for ws in (ws_test, ws_train):
        for i in range(1, 7):
            ws.column_dimensions[get_column_letter(i)].width = 26

    # Write sections stacked one after another
    row_t = 1
    for metric in metrics:
        row_t = _write_metric_section(ws_test, parsed_rows, metric, 'test', row_t)

    row_r = 1
    for metric in metrics:
        row_r = _write_metric_section(ws_train, parsed_rows, metric, 'train', row_r)

    # Ensure output directory exists
    out_dir = os.path.dirname(out_xlsx_tt)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_xlsx_tt)
    print(f"Saved Excel workbook (Test/Train combined) to: {out_xlsx_tt}")


def _filter_records_by_initial(records: List[Dict[str, Any]], allowed_keywords: List[str]) -> List[Dict[str, Any]]:
    """Filter raw records by matching parsed initial PC source against allowed keywords (case-insensitive substring)."""
    allowed_lower = [k.lower() for k in allowed_keywords]
    filtered: List[Dict[str, Any]] = []
    for r in records:
        initial, _, _ = parse_experiment_name(r.get('experiment_name', ''))
        low = initial.lower()
        if any(k in low for k in allowed_lower):
            filtered.append(r)
    return filtered


def build_excel_by_metric(records: List[Dict[str, Any]], out_xlsx_metrics: str):
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to output Excel. Please install it: pip install openpyxl")

    metrics: Tuple[str, ...] = ('PSNR', 'SSIM', 'LPIPS')
    parsed_rows = _parse_rows(records, metrics)

    wb = openpyxl.Workbook()
    # remove default sheet; we'll create per-metric sheets
    wb.remove(wb.active)

    for metric in metrics:
        title = metric if len(metric) <= 31 else metric[:31]
        ws = wb.create_sheet(title=title)
        # column widths
        for i in range(1, 7):
            ws.column_dimensions[get_column_letter(i)].width = 26

        # Order: Train then Test (as requested)
        row = 1
        row = _write_metric_section(ws, parsed_rows, metric, 'train', row)
        row = _write_metric_section(ws, parsed_rows, metric, 'test', row)

    out_dir = os.path.dirname(out_xlsx_metrics)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_xlsx_metrics)
    print(f"Saved Excel workbook (by metric) to: {out_xlsx_metrics}")


def main():
    parser = argparse.ArgumentParser(description='Create Excel report with custom metric tables')
    parser.add_argument('--dir', required=True, type=str, help='Root directory to search for results_* JSON files')
    parser.add_argument('--out_xlsx', required=True, type=str, help='Path to output Excel workbook')
    parser.add_argument('--out_xlsx_tt', required=False, type=str, help='Path to output Test/Train-combined Excel workbook; defaults to <out_xlsx basename>_TestTrain.xlsx')
    parser.add_argument('--out_xlsx_metrics', required=False, type=str, help='Path to output Metrics-grouped Excel workbook; defaults to <out_xlsx basename>_ByMetric.xlsx')
    parser.add_argument('--out_relevant', required=False, type=str, help='Base path or directory for filtered (relevant) Excel outputs: regular, _TestTrain, _ByMetric')
    args = parser.parse_args()

    records = collect_records(args.dir)
    if not records:
        print('No results found in directory:', args.dir)
        return

    build_excel(records, args.out_xlsx)
    # Derive default for TestTrain workbook if not provided
    out_tt = args.out_xlsx_tt
    if not out_tt:
        base, ext = os.path.splitext(args.out_xlsx)
        if not ext:
            ext = '.xlsx'
        out_tt = f"{base}_TestTrain{ext}"
    build_excel_test_train(records, out_tt)

    out_metrics = args.out_xlsx_metrics
    if not out_metrics:
        base, ext = os.path.splitext(args.out_xlsx)
        if not ext:
            ext = '.xlsx'
        out_metrics = f"{base}_ByMetric{ext}"
    build_excel_by_metric(records, out_metrics)

    # If requested, create filtered "relevant" variants for specific initial PC sources
    if args.out_relevant:
        relevant_keys = ['LidarAndColmapCropped', 'LidarOnly', 'ColmapCropped']
        filtered_records = _filter_records_by_initial(records, relevant_keys)
        if not filtered_records:
            print('Warning: No relevant records matched the requested filters:', ', '.join(relevant_keys))
        # Determine base output paths
        out_rel = args.out_relevant
        if out_rel.lower().endswith('.xlsx'):
            rel_base, rel_ext = os.path.splitext(out_rel)
            rel_main = out_rel
            rel_tt = f"{rel_base}_TestTrain{rel_ext}"
            rel_metrics = f"{rel_base}_ByMetric{rel_ext}"
        else:
            os.makedirs(out_rel, exist_ok=True)
            rel_main = os.path.join(out_rel, 'relevant.xlsx')
            rel_tt = os.path.join(out_rel, 'relevant_TestTrain.xlsx')
            rel_metrics = os.path.join(out_rel, 'relevant_ByMetric.xlsx')

        build_excel(filtered_records, rel_main)
        build_excel_test_train(filtered_records, rel_tt)
        build_excel_by_metric(filtered_records, rel_metrics)


if __name__ == '__main__':
    main()
